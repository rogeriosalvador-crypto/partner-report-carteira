import os, json, io
from pathlib import Path
from functools import wraps
from flask import (Flask, render_template, request, redirect,
                   url_for, session, send_file, abort, jsonify)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ────────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "partner-secret-2024")

BASE_DIR  = Path(__file__).parent
DATA_FILE = BASE_DIR / "data" / "relatorio.json"
ROG_DIR   = BASE_DIR / "static" / "rog"
GRUPOS    = ["NAGUMO", "FESTVAL", "JACOMAR"]
PERIODOS  = ["D-1", "D-7", "D-15", "D-21", "MTD", "CONSOLIDADO"]

# ── Data helpers ──────────────────────────────────────────────────────────────
def load_data(grupo):
    d = json.loads(DATA_FILE.read_text())
    return {
        "data_referencia": d["data_referencia"],
        "data_atualizacao": d["data_atualizacao"],
        "grupo": grupo,
        "periodos": d["periodos"][grupo],
        "lojas": d["lojas"][grupo],
    }

def load_users():
    return json.loads((BASE_DIR / "users.json").read_text())

# ── Auth ──────────────────────────────────────────────────────────────────────
def validate_user(grupo, usuario, senha):
    for u in load_users().get(grupo.upper(), []):
        if u["usuario"] == usuario and u["senha"] == senha:
            return True
    return False

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("usuario"):
            g = kwargs.get("grupo", "nagumo")
            return redirect(url_for("login_page", grupo=g))
        return f(*args, **kwargs)
    return decorated

# ── Formatting ────────────────────────────────────────────────────────────────
def fmt_brl(v):
    try:
        return "R$ {:,.2f}".format(float(v)).replace(",","X").replace(".",",").replace("X",".")
    except Exception:
        return "—"

def fmt_pct(v, d=1):
    try: return f"{float(v):.{d}f}%"
    except Exception: return "—"

def fmt_num(v):
    try: return f"{int(float(v)):,}".replace(",",".")
    except Exception: return "—"

def semaforo(campo, val):
    if val is None: return ""
    try: v = float(val)
    except Exception: return ""
    rules = {
        "ating_pct": [(100,"verde"),(90,"amarelo")],
        "cancel_pct":     [(3,"verde"),(6,"amarelo")],
        "ruptura_pct":    [(20,"verde"),(30,"amarelo")],
        "nps":            [(75,"verde"),(60,"amarelo")],
        "sla_pct":        [(85,"verde"),(70,"amarelo")],
        "online_pct":     [(90,"verde"),(80,"amarelo")],
        "nsu_pct":        [(2,"verde"),(5,"amarelo")],
    }
    thresholds = rules.get(campo, [])
    if not thresholds: return ""
    t1, c1 = thresholds[0]
    t2, c2 = thresholds[1]
    # Metrics where lower is better (cancel, ruptura, nsu)
    if campo in ("cancel_pct","ruptura_pct","nsu_pct"):
        if v <= t1: return c1
        if v <= t2: return c2
    else:
        if v >= t1: return c1
        if v >= t2: return c2
    return "vermelho"

def loja_rows_html(lojas_list):
    rows = ""
    for lj in lojas_list:
        def sem_cell(campo, val, fn=fmt_pct):
            if val is None: return "<td>—</td>"
            cor = semaforo(campo, val)
            return f'<td class="sem-{cor}">{fn(val)}</td>'
        aov = lj.get("aov")
        rows += (
            f"<tr><td>{lj.get('nome','—')}</td>"
            f"<td>{fmt_brl(lj.get('gmv'))}</td>"
            f"<td>{fmt_num(lj.get('pedidos'))}</td>"
            f"{sem_cell('ating_pct', lj.get('ating_pct'))}"
            f"<td>{fmt_brl(aov) if aov else '—'}</td>"
            f"{sem_cell('cancel_pct', lj.get('cancel_pct'))}"
            f"{sem_cell('ruptura_pct', lj.get('ruptura_pct'))}"
            f"<td>{fmt_pct(lj.get('er_pct'))}</td>"
            f"{sem_cell('nps', lj.get('nps'), fmt_num)}"
            f"{sem_cell('sla_pct', lj.get('sla_pct'))}"
            f"{sem_cell('online_pct', lj.get('online_pct'))}"
            f"<td>{fmt_brl(lj.get('gmv_rupt'))}</td>"
            f"<td>{fmt_brl(lj.get('gmv_recup'))}</td>"
            f"{sem_cell('nsu_pct', lj.get('nsu_pct'))}</tr>"
        )
    return rows

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return redirect(url_for("login_page", grupo="nagumo"))

@app.route("/<grupo>")
def login_page(grupo):
    g = grupo.upper()
    if g not in GRUPOS: abort(404)
    if session.get("usuario") and session.get("grupo","").upper() == g:
        return redirect(url_for("dashboard", grupo=g.lower()))
    return render_template("login.html", grupo=g, error=None)

@app.route("/login", methods=["POST"])
def login():
    grupo   = request.form.get("grupo","").upper()
    usuario = request.form.get("usuario","").strip().lower()
    senha   = request.form.get("senha","").strip()
    if grupo not in GRUPOS: abort(400)
    if validate_user(grupo, usuario, senha):
        session["grupo"] = grupo
        session["usuario"] = usuario
        return redirect(url_for("dashboard", grupo=grupo.lower()))
    return render_template("login.html", grupo=grupo,
                           error="Usuário ou senha incorretos.")

@app.route("/dashboard/<grupo>")
@login_required
def dashboard(grupo):
    g = grupo.upper()
    if g not in GRUPOS: abort(404)
    if session.get("grupo","").upper() != g:
        return redirect(url_for("dashboard",
                                grupo=session.get("grupo","nagumo").lower()))
    data   = load_data(g)
    p      = data["periodos"]
    d1     = p.get("D-1", {})
    mtd    = p.get("MTD", {})
    cons   = p.get("CONSOLIDADO", {})

    # KPI cards
    kpi_cards = [
        {"titulo":"GMV D-1","valor":fmt_brl(d1.get("gmv")),
         "badge":fmt_pct(d1.get("ating_pct")),
         "cor":semaforo("ating_pct", d1.get("ating_pct"))},
        {"titulo":"Pedidos","valor":fmt_num(d1.get("pedidos")),"badge":None,"cor":""},
        {"titulo":"AOV","valor":fmt_brl(d1.get("aov")),"badge":None,"cor":""},
        {"titulo":"Cancelamento","valor":fmt_pct(d1.get("cancel_pct")),
         "badge":None,"cor":semaforo("cancel_pct",d1.get("cancel_pct"))},
        {"titulo":"Ruptura","valor":fmt_pct(d1.get("ruptura_pct")),
         "badge":None,"cor":semaforo("ruptura_pct",d1.get("ruptura_pct"))},
        {"titulo":"NPS","valor":fmt_num(d1.get("nps")),
         "badge":None,"cor":semaforo("nps",d1.get("nps"))},
    ]

    # Comparativo rows
    comp_rows = []
    for label, key, fn in [
        ("GMV (R$)","gmv",fmt_brl), ("Pedidos","pedidos",fmt_num),
        ("AOV (R$)","aov",fmt_brl), ("Cancel%","cancel_pct",fmt_pct),
        ("Ruptura%","ruptura_pct",fmt_pct), ("NPS","nps",fmt_num),
    ]:
        row = [label]
        for per in ["D-1","D-7","D-15","D-21"]:
            v = p.get(per,{}).get(key)
            row.append(fn(v) if v is not None else "—")
        comp_rows.append(row)

    # Loja sections per period
    loja_data = {per: loja_rows_html(data["lojas"].get(per, []))
                 for per in PERIODOS}

    mtd_kpis = [
        {"label":"GMV MTD","val":fmt_brl(mtd.get("gmv"))},
        {"label":"Meta MTD","val":fmt_brl(mtd.get("meta_gmv"))},
        {"label":"Atingimento","val":fmt_pct(mtd.get("ating_pct"))},
        {"label":"Pedidos MTD","val":fmt_num(mtd.get("pedidos"))},
        {"label":"NPS MTD","val":fmt_num(mtd.get("nps"))},
    ]
    cons_kpis = [
        {"label":"GMV Realizado","val":fmt_brl(cons.get("gmv"))},
        {"label":"Meta","val":fmt_brl(cons.get("meta_gmv"))},
        {"label":"Atingimento","val":fmt_pct(cons.get("ating_pct"))},
        {"label":"Pedidos","val":fmt_num(cons.get("pedidos"))},
    ]
    return render_template("dashboard.html",
        grupo=g, data_referencia=data["data_referencia"],
        data_atualizacao=data["data_atualizacao"],
        kpi_cards=kpi_cards, comp_rows=comp_rows,
        loja_data_json=json.dumps(loja_data),
        mtd_kpis=mtd_kpis, cons_kpis=cons_kpis)

@app.route("/atualizar", methods=["POST"])
def atualizar():
    payload = request.get_json(silent=True) or {}
    if payload:
        DATA_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    return jsonify({"ok": True})

@app.route("/api/dados/<grupo>")
def api_dados(grupo):
    g = grupo.upper()
    if g not in GRUPOS: abort(404)
    return jsonify(load_data(g))

@app.route("/download/excel/<grupo>")
@login_required
def download_excel(grupo):
    g = grupo.upper()
    if g not in GRUPOS: abort(404)
    if session.get("grupo","").upper() != g: abort(403)
    data = load_data(g)
    buf = io.BytesIO()
    _build_excel(data).save(buf)
    buf.seek(0)
    fname = f"Relatorio_{g}_{data['data_referencia'].replace('/','')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/logout")
def logout():
    g = session.get("grupo","nagumo").lower()
    session.clear()
    return redirect(url_for("login_page", grupo=g))

# ── Excel builder ─────────────────────────────────────────────────────────────
def _hdr(ws, row, ncols, color="C62828"):
    fill = PatternFill("solid", fgColor=color)
    fnt  = Font(bold=True, color="FFFFFF", size=10)
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = fnt
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _build_excel(data):
    wb = Workbook(); grupo = data["grupo"]; ref = data["data_referencia"]
    p = data["periodos"]; lojas = data["lojas"]
    # Resumo
    ws1 = wb.active; ws1.title = "Resumo"
    ws1.merge_cells("A1:F1"); ws1["A1"].value = f"Resumo — {grupo} (ref: {ref})"
    ws1["A1"].font = Font(bold=True, size=13, color="C62828")
    ws1["A1"].alignment = Alignment(horizontal="center")
    for i,h in enumerate(["Indicador","D-1","D-7","D-15","D-21","MTD"],1):
        ws1.cell(row=2,column=i).value = h
    _hdr(ws1, 2, 6)
    for ri,(label,key,fn) in enumerate([
        ("GMV (R$)","gmv",fmt_brl),("Pedidos","pedidos",fmt_num),
        ("AOV (R$)","aov",fmt_brl),("Cancel%","cancel_pct",fmt_pct),
        ("Ruptura%","ruptura_pct",fmt_pct),("NPS","nps",fmt_num),
        ("Ating%","ating_pct",fmt_pct),
    ], 3):
        ws1.cell(row=ri,column=1).value = label
        ws1.cell(row=ri,column=1).font = Font(bold=True)
        for ci,per in enumerate(["D-1","D-7","D-15","D-21","MTD"],2):
            v = p.get(per,{}).get(key)
            ws1.cell(row=ri,column=ci).value = fn(v) if v is not None else "—"
    for i,ltr in enumerate(['A','B','C','D','E','F'],1): ws1.column_dimensions[ltr].width = 18
    # Lojas D-1
    ws2 = wb.create_sheet("Lojas D-1")
    hdrs2 = ["Loja","GMV","Pedidos","Ating%","Cancel%","Ruptura%","NPS","SLA%","Online%","NSU%"]
    for i,h in enumerate(hdrs2,1): ws2.cell(row=1,column=i).value = h
    _hdr(ws2, 1, len(hdrs2))
    for ri,lj in enumerate(lojas.get("D-1",[]),2):
        for ci,val in enumerate([
            lj.get("nome",""), fmt_brl(lj.get("gmv")), fmt_num(lj.get("pedidos")),
            fmt_pct(lj.get("ating_pct")), fmt_pct(lj.get("cancel_pct")),
            fmt_pct(lj.get("ruptura_pct")), fmt_num(lj.get("nps")),
            fmt_pct(lj.get("sla_pct")), fmt_pct(lj.get("online_pct")),
            fmt_pct(lj.get("nsu_pct")),
        ],1): ws2.cell(row=ri,column=ci).value = val
    ws2.column_dimensions["A"].width = 40
    for col in list(ws2.columns)[1:]: ws2.column_dimensions[col[0].column_letter].width = 14
    # MTD
    ws3 = wb.create_sheet("MTD")
    hdrs3 = ["Loja","GMV","Pedidos","Ating%","Cancel%","Ruptura%","NPS","NSU%"]
    for i,h in enumerate(hdrs3,1): ws3.cell(row=1,column=i).value = h
    _hdr(ws3, 1, len(hdrs3))
    for ri,lj in enumerate(lojas.get("MTD",[]),2):
        for ci,val in enumerate([
            lj.get("nome",""), fmt_brl(lj.get("gmv")), fmt_num(lj.get("pedidos")),
            fmt_pct(lj.get("ating_pct")), fmt_pct(lj.get("cancel_pct")),
            fmt_pct(lj.get("ruptura_pct")), fmt_num(lj.get("nps")),
            fmt_pct(lj.get("nsu_pct")),
        ],1): ws3.cell(row=ri,column=ci).value = val
    ws3.column_dimensions["A"].width = 40
    for col in list(ws3.columns)[1:]: ws3.column_dimensions[col[0].column_letter].width = 14
    return wb


@app.route("/health")
def health():
    return jsonify({"status": "ok"}), 200


# ── Entrypoint ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)), debug=False)
